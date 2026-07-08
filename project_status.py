import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ── Styles ──
header_font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
complete_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
pending_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
partial_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
section_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
section_font = Font(name="Calibri", bold=True, size=11, color="2F5496")
normal_font = Font(name="Calibri", size=11)
bold_font = Font(name="Calibri", bold=True, size=11)
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
wrap_align = Alignment(wrap_text=True, vertical="top")

def style_header(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

def style_row(ws, row, cols, status=None):
    fill = None
    if status == "Done":
        fill = complete_fill
    elif status == "Pending":
        fill = pending_fill
    elif status == "Partial":
        fill = partial_fill
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = normal_font
        cell.border = thin_border
        cell.alignment = wrap_align
        if fill:
            cell.fill = fill

def section_row(ws, row, cols, title):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = section_font
    cell.fill = section_fill
    cell.alignment = Alignment(vertical="center")
    for c in range(1, cols + 1):
        ws.cell(row=row, column=c).border = thin_border
        ws.cell(row=row, column=c).fill = section_fill

# ═══════════════════════════════════════════════════════════════
# SHEET 1: Project Status Overview
# ═══════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Project Status"
ws1.sheet_properties.tabColor = "2F5496"

headers = ["#", "Module / Feature", "Sub-Feature", "Status", "Priority", "Notes"]
cols = len(headers)
ws1.append(headers)
style_header(ws1, 1, cols)

data = [
    # ── Core Interview Engine ──
    ("CORE INTERVIEW ENGINE", None),
    (1, "Resume Parsing", "PDF/DOCX upload & extraction", "Done", "P0", "pdfplumber + docx2txt"),
    (2, "Resume Parsing", "Candidate profiling (domain, level, skills)", "Done", "P0", "Auto-detects from resume"),
    (3, "Question Generation", "LLM-powered adaptive questions", "Done", "P0", "Uses conversation history"),
    (4, "Question Generation", "Follow-up question logic", "Done", "P0", "Based on answer quality"),
    (5, "Question Generation", "Contradiction pairs testing", "Done", "P1", "3 pairs per domain/level"),
    (6, "Answer Processing", "Real-time STT transcription", "Done", "P0", "OpenAI Whisper, Deepgram, Inworld"),
    (7, "Answer Processing", "TTS response synthesis", "Done", "P0", "Deepgram Aura, Inworld, OpenAI, Kugel"),
    (8, "Evaluation", "Per-question scoring", "Done", "P0", "Structured rubric per level"),
    (9, "Evaluation", "Overall session evaluation", "Done", "P0", "Score, grade, trajectory, recommendation"),
    (10, "Evaluation", "Behavioral analysis", "Done", "P1", "Filler words, pauses, pronouns"),
    (11, "Session Management", "In-memory + PostgreSQL state", "Done", "P0", "Dual-layer with Redis cache"),
    (12, "Session Management", "Cross-worker session sharing", "Done", "P0", "DB-backed sync"),
    (13, "WebSocket Audio", "Real-time audio streaming", "Done", "P0", "Browser VAD + server processing"),

    # ── Domains & Levels ──
    ("DOMAINS & EXPERIENCE LEVELS", None),
    (14, "Domain: Physical Design", "Interview prompts", "Done", "P0", "experienced_junior + experienced_senior"),
    (15, "Domain: Analog Layout", "Interview prompts", "Done", "P0", "experienced_junior + experienced_senior"),
    (16, "Domain: Design Verification", "Interview prompts", "Done", "P0", "experienced_junior + experienced_senior"),
    (17, "Domain: Physical Design", "Eval prompts", "Done", "P0", "experienced_junior + experienced_senior"),
    (18, "Domain: Analog Layout", "Eval prompts", "Done", "P0", "experienced_junior + experienced_senior"),
    (19, "Domain: Design Verification", "Eval prompts", "Done", "P0", "experienced_junior + experienced_senior"),
    (20, "Level: fresh_graduate", "Interview + eval prompts", "Pending", "P1", "No prompts exist yet"),
    (21, "Level: trained_fresher", "Interview + eval prompts", "Pending", "P1", "Backup files exist (.bak) but not active"),
    (22, "STT Domain Hints", "Per-domain transcription prompts", "Done", "P1", "3 domain-specific hint files"),

    # ── Anti-Cheat & Proctoring ──
    ("ANTI-CHEAT & PROCTORING", None),
    (23, "Gaze Tracking", "ML ensemble classifier", "Done", "P1", "sklearn model (gaze_ensemble_model.pkl)"),
    (24, "Face Verification", "AWS Rekognition face compare", "Done", "P1", "Register + compare + liveness"),
    (25, "Face Verification", "Glasses detection", "Done", "P2", "AWS Rekognition"),
    (26, "Speaker Verification", "Resemblyzer voice matching", "Done", "P1", "256-dim embeddings, threshold 0.75"),
    (27, "Speaker Verification", "ECAPA-TDNN model integration", "Pending", "P2", "Model downloaded at models/spkrec-ecapa/ but NOT integrated"),
    (28, "Speaker Verification", "Mismatch counting (no interview end)", "Done", "P1", "Counts mismatches, reports in callback"),
    (29, "Tab Switch Detection", "Browser tab change logging", "Done", "P1", "Logged in anticheat_log"),
    (30, "AI Detection", "AI-generated answer flagging", "Done", "P2", "Flags in anticheat_log"),
    (31, "Keystroke Monitoring", "Copy-paste & keystroke events", "Done", "P2", "Behavioral event logging"),

    # ── LMS Integration ──
    ("LMS INTEGRATION", None),
    (32, "LMS Launch", "POST /api/lms/launch endpoint", "Done", "P0", "SAML/OAuth, resume upload, redirect"),
    (33, "LMS Callback", "POST results to callback_url", "Done", "P0", "Structured payload with Q&A, retry logic"),
    (34, "LMS Callback", "Student + result + questions + integrity", "Done", "P0", "Full transcript included"),
    (35, "LMS Callback", "Staging endpoint tested", "Done", "P1", "http://13.232.251.185 — 200 OK"),
    (36, "LMS Callback", "Production endpoint", "Pending", "P0", "eduspark.sumedhait.com returns 404 — LMS team needs to deploy"),
    (37, "LMS Integration Guide", "Documentation for LMS team", "Done", "P1", "LMS_INTEGRATION_GUIDE.md"),
    (38, "LMS Test Page", "lms_test.html testing UI", "Done", "P2", "SAML/OAuth launch simulation"),

    # ── Admin Dashboard ──
    ("ADMIN DASHBOARD", None),
    (39, "Session Browser", "List/filter/view sessions", "Done", "P0", "With search, date range, domain filter"),
    (40, "LLM Config", "Model switcher (17 models)", "Done", "P0", "qgen + eval model selection"),
    (41, "Voice Config", "TTS/STT provider switcher", "Done", "P1", "4 TTS + 3 STT providers"),
    (42, "Prompt Editor", "Edit eval prompts in-browser", "Done", "P1", "Save/reset per level"),
    (43, "Expert Reviews", "Submit human expert scores", "Done", "P1", "Real INSERT + auto AI score lookup"),
    (44, "Share Links", "Public session review URLs", "Done", "P2", "Token-based, no auth required"),
    (45, "Rerun Evaluation", "Re-evaluate ended sessions", "Done", "P1", "POST /api/admin/rerun-eval/{sid}"),
    (46, "Prompt Playground", "Test eval prompts live", "Done", "P2", "POST /api/admin/prompt-playground"),
    (47, "Anticheat Config", "Configure thresholds", "Done", "P2", "GET/POST /api/admin/anticheat-config"),

    # ── Cognition AI (Self-Improving) ──
    ("COGNITION AI (SELF-IMPROVING ENGINE)", None),
    (48, "Signal Collectors", "eval_parse_fail detector", "Done", "P1", "Detects evaluation parse errors"),
    (49, "Signal Collectors", "difficulty_ramp detector", "Done", "P1", "Score drop >20% mid-interview"),
    (50, "Signal Collectors", "topic_gap detector", "Done", "P1", "Resume skills never asked about"),
    (51, "Signal Collectors", "cost_anomaly detector", "Done", "P1", "Token usage >2σ from mean"),
    (52, "Signal Collectors", "followup_quality detector", "Done", "P1", "Follow-ups scoring 25%+ lower"),
    (53, "Signal Collectors", "score_drift detector", "Done", "P1", "Expert vs AI disagreement >1.5"),
    (54, "Diagnosis Engine", "LLM-powered prompt fix suggestions", "Done", "P1", "Auto-triggers at 3+ signals, 24h cooldown"),
    (55, "Background Sweeper", "600s interval incremental scan", "Done", "P1", "Watermark-based, daemon thread"),
    (56, "Admin Endpoints", "Signals/summary/diagnoses/trigger", "Done", "P1", "5 endpoints under /api/admin/cognition/"),
    (57, "Cognition Model Config", "Separate cognition_model in RUNTIME_CONFIG", "Done", "P2", "us.anthropic.claude-haiku"),
    (58, "Auto-Apply Fixes", "Apply diagnosis suggestions to prompts", "Pending", "P2", "Manual review only — no auto-patching yet"),

    # ── Cost & Observability ──
    ("COST & OBSERVABILITY", None),
    (59, "LLM Cost Tracking", "Per-call token + cost logging", "Done", "P1", "17-model pricing matrix"),
    (60, "Observability Dashboard", "Summary + log query endpoints", "Done", "P1", "/api/observability/summary + /logs"),
    (61, "Load Testing", "loadtest.py (4 scenarios)", "Done", "P2", "health, session, turn, stream"),

    # ── Infrastructure & Deployment ──
    ("INFRASTRUCTURE & DEPLOYMENT", None),
    (62, "PostgreSQL", "Schema migration (15 tables)", "Done", "P0", "Auto-creates on startup"),
    (63, "Redis Cache", "Session caching (2h TTL)", "Done", "P1", "Optional — works without Redis"),
    (64, "AWS Bedrock", "Claude model access", "Done", "P0", "Primary LLM provider"),
    (65, "AWS Rekognition", "Face verification", "Done", "P1", "Register, compare, liveness"),
    (66, "Secrets Management", "Env vars + AWS SSM + Secrets Manager", "Done", "P1", "secrets_proxy.py"),
    (67, "EC2 Deployment", "nginx + uvicorn on port 8001", "Done", "P0", "Deploy via git pull"),
    (68, "EC2 Server", "Pull latest code changes", "Pending", "P0", "Server running OLD code — needs git pull + restart"),
    (69, "SSL/HTTPS", "Production HTTPS", "Partial", "P0", "nginx handles SSL but cert status unknown"),
    (70, "CI/CD Pipeline", "Automated testing & deploy", "Pending", "P2", "Manual git pull deploy currently"),
    (71, "Docker", "Containerized deployment", "Pending", "P3", "Not containerized"),
    (72, "Monitoring & Alerts", "Production health monitoring", "Pending", "P2", "No alerting system in place"),
    (73, "Logging", "Centralized log aggregation", "Pending", "P2", "Console logging only, no log aggregation"),

    # ── Testing ──
    ("TESTING", None),
    (74, "Unit Tests", "Python unit test suite", "Pending", "P2", "No test files exist"),
    (75, "Integration Tests", "API endpoint tests", "Pending", "P2", "No automated API tests"),
    (76, "Load Tests", "Performance benchmarking", "Done", "P2", "loadtest.py covers 4 scenarios"),
    (77, "Speaker Test Page", "Audio device test UI", "Done", "P2", "speaker_test.html"),

    # ── Pending / Future ──
    ("PENDING / FUTURE WORK", None),
    (78, "Voice Cloning", "Clone interviewer voice", "Pending", "P3", "Placeholder endpoint exists, not implemented"),
    (79, "Multi-Language Support", "Non-English interviews", "Pending", "P3", "English only currently"),
    (80, "Candidate Portal", "Student dashboard for past results", "Pending", "P3", "No student-facing dashboard"),
    (81, "Analytics Dashboard", "Interview analytics & trends", "Pending", "P2", "Admin sees sessions but no trend charts"),
    (82, "Email Notifications", "Send results to candidates", "Pending", "P3", "No email integration"),
    (83, "Batch Interview", "Schedule/queue multiple interviews", "Pending", "P3", "One at a time currently"),
    (84, "Mobile Responsive", "Mobile-optimized interview UI", "Pending", "P3", "Desktop-focused layout"),
    (85, "Rate Limiting", "API rate limiting", "Pending", "P2", "No rate limits on endpoints"),
]

row = 2
for item in data:
    if len(item) == 2 and item[1] is None:
        section_row(ws1, row, cols, item[0])
        row += 1
        continue
    num, module, sub, status, priority, notes = item
    ws1.cell(row=row, column=1, value=num)
    ws1.cell(row=row, column=2, value=module)
    ws1.cell(row=row, column=3, value=sub)
    ws1.cell(row=row, column=4, value=status)
    ws1.cell(row=row, column=5, value=priority)
    ws1.cell(row=row, column=6, value=notes)
    style_row(ws1, row, cols, status)
    row += 1

ws1.column_dimensions["A"].width = 5
ws1.column_dimensions["B"].width = 28
ws1.column_dimensions["C"].width = 38
ws1.column_dimensions["D"].width = 10
ws1.column_dimensions["E"].width = 10
ws1.column_dimensions["F"].width = 55

# ═══════════════════════════════════════════════════════════════
# SHEET 2: Summary Stats
# ═══════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Summary")
ws2.sheet_properties.tabColor = "548235"

ws2.merge_cells("A1:C1")
ws2.cell(row=1, column=1, value="PROJECT STATUS SUMMARY").font = Font(name="Calibri", bold=True, size=14, color="2F5496")

summary_headers = ["Metric", "Count", "Percentage"]
for c, h in enumerate(summary_headers, 1):
    cell = ws2.cell(row=3, column=c, value=h)
    cell.font = header_font
    cell.fill = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
    cell.alignment = Alignment(horizontal="center")
    cell.border = thin_border

done_count = sum(1 for d in data if len(d) == 6 and d[3] == "Done")
pending_count = sum(1 for d in data if len(d) == 6 and d[3] == "Pending")
partial_count = sum(1 for d in data if len(d) == 6 and d[3] == "Partial")
total = done_count + pending_count + partial_count

summary_data = [
    ("Total Features", total, "100%"),
    ("Completed (Done)", done_count, f"{done_count*100//total}%"),
    ("Partially Done", partial_count, f"{partial_count*100//total}%"),
    ("Pending", pending_count, f"{pending_count*100//total}%"),
    ("", "", ""),
    ("P0 (Critical)", sum(1 for d in data if len(d)==6 and d[4]=="P0"), ""),
    ("P0 Done", sum(1 for d in data if len(d)==6 and d[4]=="P0" and d[3]=="Done"), ""),
    ("P0 Pending", sum(1 for d in data if len(d)==6 and d[4]=="P0" and d[3] in ("Pending","Partial")), ""),
    ("", "", ""),
    ("P1 (Important)", sum(1 for d in data if len(d)==6 and d[4]=="P1"), ""),
    ("P1 Done", sum(1 for d in data if len(d)==6 and d[4]=="P1" and d[3]=="Done"), ""),
    ("P1 Pending", sum(1 for d in data if len(d)==6 and d[4]=="P1" and d[3]=="Pending"), ""),
    ("", "", ""),
    ("Total API Endpoints", "55+", ""),
    ("Database Tables", 15, ""),
    ("LLM Models Supported", 17, ""),
    ("TTS Providers", 4, ""),
    ("STT Providers", 3, ""),
    ("Domains", 3, ""),
    ("Experience Levels (active)", "2 of 4", ""),
    ("Scored Sessions (Production)", 12, ""),
]

for i, (metric, count, pct) in enumerate(summary_data, 4):
    ws2.cell(row=i, column=1, value=metric).font = bold_font if metric and "Total" in str(metric) or "Done" in str(metric) or "Pending" in str(metric) else normal_font
    ws2.cell(row=i, column=2, value=count).font = normal_font
    ws2.cell(row=i, column=3, value=pct).font = normal_font
    for c in range(1, 4):
        ws2.cell(row=i, column=c).border = thin_border
        ws2.cell(row=i, column=c).alignment = Alignment(horizontal="center")
    if "Done" in str(metric):
        for c in range(1, 4):
            ws2.cell(row=i, column=c).fill = complete_fill
    elif "Pending" in str(metric):
        for c in range(1, 4):
            ws2.cell(row=i, column=c).fill = pending_fill

ws2.column_dimensions["A"].width = 30
ws2.column_dimensions["B"].width = 15
ws2.column_dimensions["C"].width = 15

# ═══════════════════════════════════════════════════════════════
# SHEET 3: API Endpoints
# ═══════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("API Endpoints")
ws3.sheet_properties.tabColor = "BF8F00"

ep_headers = ["#", "Method", "Endpoint", "Category", "Purpose", "Auth Required"]
for c, h in enumerate(ep_headers, 1):
    cell = ws3.cell(row=1, column=c, value=h)
    cell.font = header_font
    cell.fill = PatternFill(start_color="BF8F00", end_color="BF8F00", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", wrap_text=True)
    cell.border = thin_border

endpoints = [
    (1, "POST", "/api/login", "Auth", "Admin login", "No"),
    (2, "POST", "/api/auth/login", "Auth", "Auth login", "No"),
    (3, "POST", "/api/auth/logout", "Auth", "Logout", "Yes"),
    (4, "GET", "/api/auth/me", "Auth", "Current user profile", "Yes"),
    (5, "POST", "/api/create-session", "Session", "Create interview session", "No"),
    (6, "GET", "/api/get-session", "Session", "Fetch session state", "No"),
    (7, "POST", "/api/start-interview", "Session", "Start interview + gen first Q", "No"),
    (8, "POST", "/api/submit-answer", "Session", "Submit answer (blocking)", "No"),
    (9, "POST", "/api/stream-answer", "Session", "Stream answer (SSE)", "No"),
    (10, "POST", "/api/end-session", "Session", "End session + auto-eval", "No"),
    (11, "POST", "/api/generate-report", "Session", "Generate eval report", "No"),
    (12, "POST", "/api/transcribe", "Speech", "STT transcription", "No"),
    (13, "WS", "/ws/audio", "Speech", "WebSocket audio streaming", "No"),
    (14, "POST", "/api/toggle-tts", "Speech", "Enable/disable TTS", "No"),
    (15, "POST", "/api/parse-resume", "Resume", "Parse PDF/DOCX resume", "No"),
    (16, "POST", "/api/lms/launch", "LMS", "Launch from LMS (SAML/OAuth)", "API Key"),
    (17, "POST", "/api/anticheat-event", "Anti-Cheat", "Log behavioral event", "No"),
    (18, "GET", "/api/anticheat-settings", "Anti-Cheat", "Get anticheat config", "No"),
    (19, "POST", "/api/anticheat/gaze", "Anti-Cheat", "Submit gaze tracking data", "No"),
    (20, "GET", "/api/face/check", "Face", "Check face reference exists", "No"),
    (21, "POST", "/api/face/register", "Face", "Register face reference", "No"),
    (22, "POST", "/api/face/detect-glasses", "Face", "Detect glasses", "No"),
    (23, "POST", "/api/face/compare", "Face", "Compare face to reference", "No"),
    (24, "GET", "/api/admin/llm-config", "Admin-LLM", "Get model config", "Yes"),
    (25, "POST", "/api/admin/llm-config", "Admin-LLM", "Switch models", "Yes"),
    (26, "GET", "/api/admin/llm-prompts", "Admin-LLM", "Get eval prompts", "Yes"),
    (27, "POST", "/api/admin/llm-prompts", "Admin-LLM", "Save/reset eval prompts", "Yes"),
    (28, "GET", "/api/admin/qgen-prompt", "Admin-LLM", "Preview qgen prompt", "Yes"),
    (29, "GET", "/api/admin/interview-prompt", "Admin-LLM", "Preview interviewer prompt", "Yes"),
    (30, "GET", "/api/tts-status", "Admin-Voice", "Check TTS status", "No"),
    (31, "GET", "/api/admin/stt-config", "Admin-Voice", "Get STT config", "Yes"),
    (32, "POST", "/api/admin/stt-config", "Admin-Voice", "Change STT provider", "Yes"),
    (33, "POST", "/api/admin/stt-test", "Admin-Voice", "Test STT with audio", "Yes"),
    (34, "GET", "/api/admin/voice-library", "Admin-Voice", "List TTS voices", "Yes"),
    (35, "POST", "/api/admin/test-tts", "Admin-Voice", "Test TTS synthesis", "Yes"),
    (36, "POST", "/api/admin/set-interview-voice", "Admin-Voice", "Set interviewer voice", "Yes"),
    (37, "POST", "/api/playground/tts", "Admin-Voice", "TTS playground", "Yes"),
    (38, "GET", "/api/admin/sessions", "Admin-Sessions", "List all sessions", "Yes"),
    (39, "GET", "/api/admin/session/{sid}", "Admin-Sessions", "Session detail", "Yes"),
    (40, "POST", "/api/admin/rerun-eval/{sid}", "Admin-Sessions", "Rerun evaluation", "Yes"),
    (41, "POST", "/api/admin/share-link", "Admin-Sessions", "Generate share link", "Yes"),
    (42, "GET", "/api/shared/session/{token}", "Public", "Shared session review", "No"),
    (43, "GET", "/api/admin/anticheat-config", "Admin-Anticheat", "Get anticheat thresholds", "Yes"),
    (44, "POST", "/api/admin/anticheat-config", "Admin-Anticheat", "Update anticheat config", "Yes"),
    (45, "POST", "/api/admin/review", "Admin-Review", "Submit expert review", "Yes"),
    (46, "GET", "/api/admin/reviews", "Admin-Review", "List all reviews", "Yes"),
    (47, "GET", "/api/admin/reviews/{sid}", "Admin-Review", "Reviews per session", "Yes"),
    (48, "GET", "/api/admin/cognition/signals", "Cognition", "List signals", "Yes"),
    (49, "GET", "/api/admin/cognition/summary", "Cognition", "Dashboard summary", "Yes"),
    (50, "GET", "/api/admin/cognition/diagnoses", "Cognition", "List diagnoses", "Yes"),
    (51, "POST", "/api/admin/cognition/diagnoses/{id}", "Cognition", "Update diagnosis status", "Yes"),
    (52, "POST", "/api/admin/cognition/trigger", "Cognition", "Manual sweep trigger", "Yes"),
    (53, "POST", "/api/admin/prompt-playground", "Admin-Playground", "Test eval prompt live", "Yes"),
    (54, "GET", "/health", "System", "Health check", "No"),
    (55, "GET", "/api/lobby-config", "System", "Available domains + levels", "No"),
    (56, "GET", "/api/domains", "System", "Supported domains", "No"),
    (57, "GET", "/api/observability/summary", "Observability", "Cost/token summary", "Yes"),
    (58, "GET", "/api/observability/logs", "Observability", "LLM call history", "Yes"),
]

for i, ep in enumerate(endpoints):
    r = i + 2
    for c, val in enumerate(ep, 1):
        ws3.cell(row=r, column=c, value=val).font = normal_font
        ws3.cell(row=r, column=c).border = thin_border
        ws3.cell(row=r, column=c).alignment = wrap_align

ws3.column_dimensions["A"].width = 5
ws3.column_dimensions["B"].width = 8
ws3.column_dimensions["C"].width = 42
ws3.column_dimensions["D"].width = 16
ws3.column_dimensions["E"].width = 35
ws3.column_dimensions["F"].width = 12

# ═══════════════════════════════════════════════════════════════
# SHEET 4: Database Schema
# ═══════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Database Schema")
ws4.sheet_properties.tabColor = "7030A0"

db_headers = ["#", "Table Name", "Purpose", "Key Columns", "Status"]
for c, h in enumerate(db_headers, 1):
    cell = ws4.cell(row=1, column=c, value=h)
    cell.font = header_font
    cell.fill = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", wrap_text=True)
    cell.border = thin_border

tables = [
    (1, "active_sessions", "Live session state (primary)", "session_id (TEXT PK), session_data (JSONB)", "Active"),
    (2, "candidate_history", "Email-indexed session history", "email, session_id, session_summary (JSONB)", "Active"),
    (3, "app_config", "Shared runtime config", "key (TEXT PK), value (JSONB)", "Active"),
    (4, "face_references", "Face encodings for verification", "email (PK), face_encoding, liveness_confidence", "Active"),
    (5, "cognition_signals", "Detected prompt quality issues", "signal_type, severity, domain, evidence (JSONB)", "Active"),
    (6, "cognition_diagnoses", "LLM prompt fix suggestions", "domain, level, problem, suggestion, status", "Active"),
    (7, "cognition_state", "Sweep watermarks", "state_key (TEXT PK), state_value (JSONB)", "Active"),
    (8, "expert_reviews", "Human expert scores", "session_id, reviewer_name, score, ai_score, delta", "Active"),
    (9, "llm_calls", "LLM cost tracking", "step, model, tokens, latency, cost_usd", "Active"),
    (10, "candidates", "Candidate profiles", "name, domain, level, skills, tools", "Legacy"),
    (11, "sessions", "Normalized sessions", "mode, domain, level, score, grade", "Legacy (unused)"),
    (12, "turns", "Q&A exchanges", "question, answer, duration, difficulty", "Legacy"),
    (13, "evaluations", "Per-turn scores", "score, quality, accuracy, expected_points", "Legacy"),
    (14, "behavioral_signals", "Speech metrics", "filler_rate, pronoun_rate, pause_duration", "Legacy"),
    (15, "reports", "Session summaries", "technical/theory/communication scores", "Legacy"),
]

for i, t in enumerate(tables):
    r = i + 2
    for c, val in enumerate(t, 1):
        ws4.cell(row=r, column=c, value=val).font = normal_font
        ws4.cell(row=r, column=c).border = thin_border
        ws4.cell(row=r, column=c).alignment = wrap_align
    status = t[4]
    fill = complete_fill if status == "Active" else PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    for c in range(1, 6):
        ws4.cell(row=r, column=c).fill = fill

ws4.column_dimensions["A"].width = 5
ws4.column_dimensions["B"].width = 22
ws4.column_dimensions["C"].width = 35
ws4.column_dimensions["D"].width = 50
ws4.column_dimensions["E"].width = 12

# ═══════════════════════════════════════════════════════════════
# SHEET 5: Action Items (Pending Work)
# ═══════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("Action Items")
ws5.sheet_properties.tabColor = "FF0000"

ai_headers = ["#", "Action Item", "Category", "Priority", "Effort", "Details"]
for c, h in enumerate(ai_headers, 1):
    cell = ws5.cell(row=1, column=c, value=h)
    cell.font = header_font
    cell.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", wrap_text=True)
    cell.border = thin_border

actions = [
    (1, "EC2: git pull + restart", "Deployment", "P0", "15 min", "Server running old code. Cognition AI, speaker mismatch counting, enhanced LMS callback not deployed yet"),
    (2, "Production callback endpoint", "LMS", "P0", "LMS team", "eduspark.sumedhait.com/api/interview-result returns 404. LMS team needs to deploy their callback route"),
    (3, "fresh_graduate prompts", "Content", "P1", "2-3 hours", "Create interview + eval prompts for fresh_graduate level (all 3 domains)"),
    (4, "trained_fresher prompts", "Content", "P1", "2-3 hours", "Restore/create interview + eval prompts for trained_fresher level (backup files exist)"),
    (5, "ECAPA-TDNN speaker verification", "Anti-Cheat", "P2", "4-6 hours", "Integrate SpeechBrain ECAPA-TDNN model (192-dim, better accuracy) to replace Resemblyzer"),
    (6, "Cognition auto-apply fixes", "Cognition AI", "P2", "3-4 hours", "Currently diagnosis suggestions are manual. Auto-apply with human approval workflow"),
    (7, "Unit test suite", "Testing", "P2", "1-2 days", "No test files exist. Add pytest tests for database.py, cognition.py, core endpoints"),
    (8, "API integration tests", "Testing", "P2", "1 day", "Automated endpoint tests with test DB"),
    (9, "Rate limiting", "Security", "P2", "2-3 hours", "No rate limits on any endpoint. Add per-IP limits on auth, session creation"),
    (10, "Monitoring & alerts", "Infrastructure", "P2", "4-6 hours", "No health monitoring or alerting. Add uptime checks, error rate alerts"),
    (11, "Log aggregation", "Infrastructure", "P2", "3-4 hours", "Console logging only. Set up CloudWatch or similar"),
    (12, "SSL cert verification", "Infrastructure", "P1", "30 min", "Verify SSL cert status and auto-renewal on production"),
    (13, "Analytics dashboard", "Admin", "P2", "1-2 days", "Interview trends, score distribution, domain breakdown charts"),
    (14, "Voice cloning", "Speech", "P3", "1-2 days", "Placeholder endpoint exists but not implemented"),
    (15, "CI/CD pipeline", "DevOps", "P2", "1 day", "Currently manual git pull. Set up GitHub Actions or similar"),
    (16, "Docker containerization", "DevOps", "P3", "4-6 hours", "Not containerized. Add Dockerfile + docker-compose"),
    (17, "Mobile responsive UI", "Frontend", "P3", "2-3 days", "Desktop-focused layout, needs mobile optimization"),
    (18, "Candidate portal", "Frontend", "P3", "2-3 days", "No student-facing dashboard for past results"),
    (19, "Email notifications", "Integration", "P3", "4-6 hours", "Send interview results to candidates via email"),
]

for i, a in enumerate(actions):
    r = i + 2
    for c, val in enumerate(a, 1):
        ws5.cell(row=r, column=c, value=val).font = normal_font
        ws5.cell(row=r, column=c).border = thin_border
        ws5.cell(row=r, column=c).alignment = wrap_align
    priority = a[3]
    if priority == "P0":
        for c in range(1, 7):
            ws5.cell(row=r, column=c).fill = pending_fill
    elif priority == "P1":
        for c in range(1, 7):
            ws5.cell(row=r, column=c).fill = partial_fill

ws5.column_dimensions["A"].width = 5
ws5.column_dimensions["B"].width = 35
ws5.column_dimensions["C"].width = 15
ws5.column_dimensions["D"].width = 10
ws5.column_dimensions["E"].width = 12
ws5.column_dimensions["F"].width = 70

# ── Save ──
output = "/home/surya/surya_project/simple_interview/Simple_Interview_Project_Status.xlsx"
wb.save(output)
print(f"Excel saved: {output}")
print(f"  Sheet 1: Project Status ({sum(1 for d in data if len(d)==6)} features)")
print(f"  Sheet 2: Summary ({done_count} done, {pending_count} pending, {partial_count} partial)")
print(f"  Sheet 3: API Endpoints ({len(endpoints)} endpoints)")
print(f"  Sheet 4: Database Schema ({len(tables)} tables)")
print(f"  Sheet 5: Action Items ({len(actions)} pending items)")
